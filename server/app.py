"""
Amazon ASIN 采集系统 v3 - Server
轻量级 FastAPI 服务器，适合 1C/2GB 低配部署
"""
import os
import io
import csv
import json
import re
import asyncio
import ipaddress
import logging
import socket
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Any, Set
from urllib.parse import urlparse

import httpx
from fastapi import FastAPI, Request, UploadFile, File, Form, Query, HTTPException
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import openpyxl

from common import config
from common.core.idents import ASIN_RE as _ASIN_RE
from common.core.zipcode import _zfill_short_numeric
from common.core.timeutil import now_ts, ts_from, utc_now
from common.database import Database
from common.dbfactory import create_database

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ==================== 全局状态 ====================

db: Optional[Database] = None
_worker_registry: Dict[str, Dict] = {}
_WORKER_ID_RE = re.compile(r'^[\w\-]{1,64}$')

# 运行时设置
_runtime_settings: dict = {}
_settings_version: int = 0
_SETTINGS_FILE = os.path.join(config.PROJECT_DIR, "runtime_settings.json")

# 全局并发协调
_global_coordinator: Dict[str, Dict] = {}

# Worker 重启标记
_worker_restart_flags: Dict[str, bool] = {}


def _default_settings() -> dict:
    return {
        "zip_code": config.DEFAULT_ZIP_CODE,
        "max_retries": config.MAX_RETRIES,
        "request_timeout": config.REQUEST_TIMEOUT,
        "session_rotate_every": config.SESSION_ROTATE_EVERY,
        "proxy_url": config.PROXY_URL,
        "token_bucket_rate": config.TOKEN_BUCKET_RATE,
        "initial_concurrency": config.INITIAL_CONCURRENCY,
        "max_concurrency": config.MAX_CONCURRENCY,
        "min_concurrency": config.MIN_CONCURRENCY,
        "global_max_concurrency": config.GLOBAL_MAX_CONCURRENCY,
        "global_max_qps": config.GLOBAL_MAX_QPS,
        "adjust_interval": config.ADJUST_INTERVAL_S,
        "target_latency": config.TARGET_LATENCY_S,
        "max_latency": config.MAX_LATENCY_S,
        "target_success_rate": config.TARGET_SUCCESS_RATE,
        "min_success_rate": config.MIN_SUCCESS_RATE,
        "block_rate_threshold": config.BLOCK_RATE_THRESHOLD,
        "cooldown_after_block": 15,
        "proxy_bandwidth_mbps": config.PROXY_BANDWIDTH_MBPS,
        "screenshot_browsers": 1,
        "screenshot_pages_per_browser": 4,
        "auto_scrape_schedules": [],
        # 后台自动重试失败任务（AI 自动化场景：不需要人工点击重试）
        "auto_retry_failed_enabled": True,
        "auto_retry_cycles": 2,          # 最多重试多少轮（每轮会再走 MAX_RETRIES 次常规重试）
        "auto_retry_delay_minutes": 5,   # 任务失败后至少等这么久才会被自动重入队
    }


def _load_settings():
    global _runtime_settings, _settings_version
    _runtime_settings = _default_settings()
    if os.path.isfile(_SETTINGS_FILE):
        try:
            with open(_SETTINGS_FILE) as f:
                saved = json.load(f)
            _runtime_settings.update(saved)
        except Exception as e:
            logger.warning(f"加载设置失败: {e}")
    _settings_version = 0


# 敏感字段：不持久化到磁盘，避免凭据泄漏（从环境变量/.env 读取）
_SENSITIVE_SETTINGS_KEYS = {"proxy_url", "tunnel_proxy_url"}


def _save_settings():
    try:
        # 剥离敏感字段后再写文件；运行时内存里仍保留，供 worker 同步
        safe = {k: v for k, v in _runtime_settings.items() if k not in _SENSITIVE_SETTINGS_KEYS}
        with open(_SETTINGS_FILE, "w") as f:
            json.dump(safe, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"保存设置失败: {e}")


def _remove_screenshot_files(file_paths: list):
    """删除截图物理文件，忽略不存在的"""
    for fp in file_paths:
        if not fp:
            continue
        # file_path 可能是相对路径 (/static/screenshots/...) 或绝对路径
        if fp.startswith("/static/"):
            full = os.path.join(config.PROJECT_DIR, "server", fp.lstrip("/"))
        else:
            full = fp
        try:
            if os.path.isfile(full):
                os.remove(full)
        except OSError:
            pass


async def _rollback_quietly(conn):
    """裸 ``BEGIN`` 块失败后回滚，两个后端共用（不按后端分叉，见 OWNERSHIP.md D-4）。

    本文件（及 Phase 3.5-3.7 从它拆出的 server/api/*.py）里曾有 7 个
    ``async with db._write_lock: BEGIN ... COMMIT`` 块。原来其中 6 个
    没有回滚路径——SQLite 下这些语句基本不会失败，失败了也只是把那一条连接留在事务里。
    PostgreSQL 下 ``db._db`` 同样是**一条**专用写连接，但事务是粘在连接上的状态：
    任何一条语句报错，事务立刻 abort，写锁随异常释放，而垫片的事务槽还是满的，
    于是**之后每一次 BEGIN** 都撞上"嵌套 BEGIN"守卫——一个失败请求把整条写路径
    永久焊死（读仍然 200，健康检查看不出来）。加回滚把它降级成"一个请求 500"。

    - 只在**错误路径**上跑：成功路径一条语句都没多，两个后端的返回值/异常类型不变。
    - 回滚本身的异常一律吞掉（SQLite 在没有活动事务时 ROLLBACK 会报错），
      永远不能盖掉真正的原因，最后一定 ``raise`` 原异常。
    - 捕获 ``BaseException`` 而不是 ``Exception``：这些块全都在 HTTP handler 里，
      客户端断开会让 Starlette 取消请求协程，``CancelledError`` 同样必须回滚，
      否则连接照样带着事务泄漏出去（common/pgdb/batches.py 的 ``_tx()`` 已经是这个口径）。

    **Phase 3.8 之后调用点只剩一个**：``server/api/worker_queue.py`` 的
    legacy ``/api/tasks/release``（批 (5)，计划 §X.1 说那条分支该删而不是收口）。
    其余 6 处随着裸 SQL 一起收进了 db 方法，那些方法各自带同样口径的回滚。
    """
    try:
        await conn.execute("ROLLBACK")
    except BaseException:
        pass


def _register_worker(worker_id: str, enable_screenshot: bool = None, ip: str = None):
    if not _WORKER_ID_RE.match(worker_id):
        return
    now = time.time()
    if worker_id not in _worker_registry:
        _worker_registry[worker_id] = {
            "worker_id": worker_id,
            "first_seen": now,
            "last_seen": now,
            "tasks_pulled": 0,
            "results_submitted": 0,
            "enable_screenshot": True,
            "ip": ip,
        }
    _worker_registry[worker_id]["last_seen"] = now
    if enable_screenshot is not None:
        _worker_registry[worker_id]["enable_screenshot"] = enable_screenshot
    if ip is not None:
        _worker_registry[worker_id]["ip"] = ip


# ==================== 生命周期 ====================

async def _scrape_event_relay():
    """Phase 2 事件流 relay 的进程入口（**PostgreSQL 专属**）。

    SQLite 后端下整个事件流一个字节的代码都不跑：``common.database.Database``
    压根没有这些方法，这里直接返回。这比运行期 ``if is_postgres()`` 守卫更强
    —— SQLite 路径上不存在可以走错的分支。

    单例由 ``pg_try_advisory_lock`` 保证：滚动部署时第二个进程拿不到锁，
    ``start_event_relay()`` 返回 False 并安静退出（见 common/pgdb/relay.py）。

    ⚠ 异常必须在这里落日志。本协程是 ``asyncio.create_task`` 起来的，没人
    ``await`` 它的结果：不接的话，relay 起不来这件事只会以一句 GC 期的
    "Task exception was never retrieved" 出现（甚至完全不出现），而现象是
    HTTP 全部正常、outbox 无声无息地涨到撑爆磁盘。事件流停摆必须是**响的**。
    """
    from common.dbfactory import is_postgres
    if not is_postgres() or db is None or not hasattr(db, "run_event_relay"):
        return
    try:
        await db.run_event_relay()
    except asyncio.CancelledError:
        raise
    except Exception:
        logger.exception(
            "事件流 relay 异常退出——采集与 HTTP 不受影响，但 scrape_outbox "
            "会持续堆积且没有新事件产生。请检查 /api/_debug/event-stream。")


async def _stop_scrape_event_relay():
    if db is not None and hasattr(db, "stop_event_relay"):
        try:
            await db.stop_event_relay()
        except Exception:
            logger.exception("停止事件流 relay 失败（继续关闭数据库）")


@asynccontextmanager
async def lifespan(app):
    global db, _callback_send_queue
    # 存储后端由 DB_BACKEND 决定（默认 sqlite，此时与移植前完全相同：
    # 同一个 common.database.Database 类、同一个无参构造）。见 common/dbfactory.py
    db = create_database()
    await db.connect()
    _load_settings()
    os.makedirs(config.EXPORT_DIR, exist_ok=True)
    os.makedirs(config.SCREENSHOT_DIR, exist_ok=True)
    os.makedirs(_SCHEDULES_DIR, exist_ok=True)
    # 回调发送队列：maxsize 限制内存上限（高峰短期堆积也不会爆内存）
    _callback_send_queue = asyncio.Queue(maxsize=1000)
    logger.info("数据库初始化完成")
    asyncio.create_task(_timeout_task_loop())
    asyncio.create_task(_auto_scrape_scheduler())
    asyncio.create_task(_completion_watcher())
    asyncio.create_task(_callback_dispatcher())
    # 后台 WAL 维护（每 120s TRUNCATE checkpoint，防 WAL 顶 64MB checkpoint 饥饿）
    db.start_maintenance(checkpoint_interval=120)
    # 启动期 optimize 改为异步：服务先就绪、worker 先能拉任务，ANALYZE 后台慢慢做
    # （此前同步执行在 2.4GB 库上会阻塞启动 2~3 分钟）
    asyncio.create_task(db.run_startup_optimize())
    asyncio.create_task(_scrape_event_relay())
    yield
    await _stop_scrape_event_relay()
    if db:
        await db.close()
    logger.info("服务器关闭")


app = FastAPI(title="Amazon Scraper v4", version="4.0.0", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=config.STATIC_DIR), name="static")


# ==================== 全局 500：结构化 JSON + 可关联的 request_id ====================
#
# 未捕获异常今天由 Starlette 的 ServerErrorMiddleware 兜底，返回
# `text/plain` 的 "Internal Server Error"，调用方拿不到任何可机读的东西、
# 也没有能和服务端日志对上的标识。这里换成与 `server/api/sync.py:_err`
# 同形状的结构化 JSON。
#
# 两个实施陷阱（对抗验证实测给出，别按"更顺手"的写法改回去）：
#
# 1. `request_id` **不能**用 `BaseHTTPMiddleware` + contextvar 生成。
#    Starlette 的中间件栈是
#    `[ServerErrorMiddleware] + user_middleware + [ExceptionMiddleware]`，
#    ServerErrorMiddleware 在**最外层**：contextvar 在下游的赋值上游看不见，
#    结果恰恰是在 500 的那一次 `request_id` 永远是 None。
#    所以这里用**纯 ASGI 中间件**把 id 写进 `scope["state"]`——scope 是同一个
#    dict 对象、按引用共享，外层的 `Request(scope).state` 一定读得到。
#
# 2. body 里**不放** `server_time_utc` 这类逐次不同的字段：任何逐次不同的字段
#    进了黄金基线就录不出来（`_err` 里那个字段之所以能留，是因为 sync 端点
#    不在黄金里）。`request_id` 可以放，因为**黄金场景里没有任何一步返回
#    500**，这条响应永远不会进基线（2.4 追加的 14 步错误路径也全是 4xx）。
#
# body 不泄漏任何异常细节：异常全文只进日志，用同一个 `request_id` 关联。


class _RequestIDMiddleware:
    """纯 ASGI 中间件：给每个 HTTP 请求生成 `request_id` 并写进 scope。

    刻意**不是** `BaseHTTPMiddleware` 子类，理由见上方第 1 条。
    只写 scope，不碰 receive/send，因此不改任何一条正常响应的字节。
    """

    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope["type"] == "http":
            # setdefault：uvicorn/TestClient 已经放了每请求一份的 state dict，
            # 这里只往里加一个键，不覆盖。
            scope.setdefault("state", {})["request_id"] = uuid.uuid4().hex
        await self.app(scope, receive, send)


app.add_middleware(_RequestIDMiddleware)


@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: Request, exc: Exception) -> JSONResponse:
    """全局 500。返回结构化 JSON，异常细节只进日志。

    注意 `ServerErrorMiddleware` 调完本 handler **仍会 re-raise**，
    所以 `TestClient(raise_server_exceptions=True)`（黄金夹具的默认值）
    下的测试期行为不变。
    """
    request_id = getattr(request.state, "request_id", None)
    logger.exception(
        f"未处理异常 request_id={request_id} {request.method} {request.url.path}"
    )
    return JSONResponse(
        status_code=500,
        content={
            "error": "internal_error",
            "detail": "服务器内部错误",
            "request_id": request_id,
        },
    )

# Phase 3 —— catalog_sync 拉取契约（GET/POST /api/v1/sync/*）。
#
# * **无条件挂载，两个后端都挂**。SQLite 上四个端点如实回 503 而不是消失：
#   不挂 = 404，而消费者会把 404 读成"暂无数据"并静默停摆（计划 §Phase 3）。
#   路由表在 import 时定型、DB_BACKEND 是运行期变量，条件挂载必然骗人。
# * ``include_in_schema=False`` 在 router 上，所以 ``/openapi.json``（黄金基线
#   step 5，逐字节钉死）不变 —— 有用例钉住这一条。
# * import 方向是单向的：sync.py **不在模块级** import server.app
#   （那是循环导入，启动即崩、整个 erpAPI 下线），它在每个 handler 里惰性取 db。
from server.api import sync as _sync_api  # noqa: E402

app.include_router(_sync_api.router)

# /api/export/incremental 的 router **不在这里挂**（Phase 3.7 顺序局部化）。
# 它落在 /api/export/{batch_name} 这条 catch-all 的前缀里，Starlette 按注册
# 顺序匹配，挪到 catch-all 之后就静默退化成 404「批次不存在: incremental」。
# 这条依赖过去是跨文件的（顶部 include、几百行之后才定义 catch-all），现在
# 由 server/api/export.py 在自己第一个 @router.get 之前 include 增量 router
# 来保证 —— 本文件的 include 列表怎么重排都打不破它。


# 中国时间（UTC+8）：用于批次名/导出文件名/定时任务调度判定，与前端展示口径一致。
# 注意：数据库里存储的 created_at/updated_at 仍是 UTC（系统时区 Etc/UTC），由前端 +8 显示。
_CN_TZ_OFFSET = timedelta(hours=8)


def _cn_now():
    """当前中国时间（Asia/Shanghai, UTC+8）。仅用于命名与调度判定，不用于 DB 存储。

    ⚠ **它不是 ``common.core.timeutil.now_ts()``，两者永远不该互换。**
    ``now_ts()`` 是**落库**的 UTC 时间戳（created_at / updated_at 那一族，
    格式契约见 timeutil 的 docstring）；``_cn_now()`` 是**展示与调度**口径的
    UTC+8 naive datetime，只喂批次名 / 导出文件名 / ``last_run_date`` 比对。
    互换的后果：拿 now_ts() 命名 → 批次名和定时任务触发点整体偏 8 小时；
    拿 _cn_now() 落库 → 往 UTC 列里灌 UTC+8，正是 Phase 4.3 修掉的那类错位。
    Phase 4.3 因此**刻意不动这个函数**。
    """
    return datetime.utcnow() + _CN_TZ_OFFSET


def _batch_name(prefix: str) -> str:
    """批次名的唯一构造点：``<prefix>_YYYYMMDD_HHMMSS``（中国时间）。

    ------------------------------------------------------------------
    P4.7 —— 为什么要收一份，以及为什么统一到**秒**精度
    ------------------------------------------------------------------
    收口之前有五处各写各的 f-string，其中 ``auto_*`` 那两处
    （``_auto_scrape_scheduler`` 与 ``/api/schedules/{name}/run``）
    是**分钟**精度，另外三处是秒。

    「同名批次不过是个 no-op」这个直觉是**错的**，这是本条真正要修的东西：
    ``common/database.py:create_batch`` 是 ``INSERT OR IGNORE`` 之后
    ``SELECT id FROM batches WHERE name=?``，撞名时返回的是**既有批次的 id**；
    随后 ``create_tasks`` 才靠 tasks 上的 ``INSERT OR IGNORE`` 吃掉重复 ASIN。

    于是「同一分钟内自动触发 + 手动触发同一个定时任务 = 什么都不做」
    **只在 ASIN 清单没变时成立**。清单变了（文件加了行、或者走全库而库里
    新增了 ASIN），新 ASIN 会被**悄悄塞进上一个批次**——批次的语义
    「一次采集」就此破掉，而且没有任何一侧会报错。

    ⚠ **本条今天只对自动调度那两条路径成立了，但它们恰好就是本函数的用户。**
    ``POST /api/upload`` 已经改成撞名 -> 409（走 ``create_batch_if_absent``），
    那条路上的静默合并不复存在；而 ``_auto_scrape_scheduler`` 与
    ``/api/schedules/{name}/run`` 仍然走 ``create_batch``（撞名返回既有 id，
    静默合并照旧）—— 它们是内部触发、没有调用方接得住 409，让定时任务因为
    撞名而整轮不跑是更坏的结果。所以**秒精度对它们仍然是承重的**，
    ``tests/test_batch_name_precision.py`` 那两条守卫一条都不能撤。

    **有意的行为改动，已声明**：``auto_*`` 两处从分钟精度升到秒精度。
    同一分钟内的第二次触发从此建**新批次**，而不是往上一个里塞。

    黄金基线不受影响（已自行复核，不是照抄结论）：
      * ``tests/golden/harness.py`` 的 Recorder 只记 status / content_type /
        body，**不记任何 header** —— ``/api/export/all`` 的批次名只出现在
        ``Content-Disposition`` 里，body 里没有；
      * 基线的 ``export_all_csv`` body 实测不含任何批次名；
      * 场景里三次 ``/api/upload`` 全部显式传 ``batch_name``
        （``BATCH_A`` / ``BATCH_B`` / ``golden_batch_*``），走不到默认分支；
      * 场景不打 ``/api/upload-sellers``，也不打
        ``/api/schedules/{name}/run``；自动调度是后台协程，夹具已 no-op。

    ⚠ 用 ``_cn_now()``（UTC+8）而不是 ``common.core.timeutil.now_ts()``：
    批次名是**展示与调度**口径。理由见 ``_cn_now()`` 的 docstring。
    ⚠ 时间戳在**调用时**现取。调用方手里就算已经有一个 ``now``，也不要
    把它传进来 —— 那会让「批次名」和「调度判定」耦合成一个参数，
    而 4.7 之后 ``_cn_now()`` 的调用点从 2 处涨到 5 处这件事本身已经登记在案
    （``_cn_now()`` 在非 UTC 主机上的口径问题是 4.3 §6 的未决项，本轮不动）。
    """
    return f"{prefix}_{_cn_now():%Y%m%d_%H%M%S}"


# ==================== 后台任务 ====================

async def _timeout_task_loop():
    """每 30s 心跳感知回收 + 硬超时兜底 + 清理离线 worker + WAL checkpoint"""
    _wal_checkpoint_counter = 0
    while True:
        await asyncio.sleep(30)
        try:
            # 识别死 Worker：2 分钟无心跳
            now = time.time()
            heartbeat_cutoff = now - 120
            dead_worker_ids = [
                wid for wid, w in _worker_registry.items()
                if w["last_seen"] < heartbeat_cutoff
            ]

            # 回收死 Worker 的 processing 任务 + 硬超时兜底（合成一条 SQL 防双重 epoch bump）
            await db.reclaim_dead_worker_tasks(dead_worker_ids)

            # 自动重试失败任务（AI 场景免人工点击）
            if _runtime_settings.get("auto_retry_failed_enabled", True):
                try:
                    await db.auto_retry_failed_tasks(
                        max_auto_cycles=int(_runtime_settings.get("auto_retry_cycles", 2)),
                        delay_minutes=int(_runtime_settings.get("auto_retry_delay_minutes", 5)),
                    )
                except Exception as e:
                    logger.warning(f"auto_retry_failed_tasks 异常: {e}")

            # 清理 5 分钟无心跳的 worker 注册信息
            offline_cutoff = now - 300
            truly_dead = [wid for wid, w in _worker_registry.items() if w["last_seen"] < offline_cutoff]
            for wid in truly_dead:
                del _worker_registry[wid]
                _global_coordinator.pop(wid, None)

            # 清理导出临时文件（超过 1 小时）：
            #   export_*.xlsx      —— xlsx 流式导出的临时文件
            #   screenshots_*.zip  —— 截图 ZIP 流式导出的临时文件
            # 正常路径由各自的 finally 删除；此处兜底进程崩溃/异常退出遗留的孤儿文件。
            try:
                export_dir = config.EXPORT_DIR
                if os.path.isdir(export_dir):
                    for fname in os.listdir(export_dir):
                        is_export_tmp = fname.startswith("export_") and fname.endswith(".xlsx")
                        is_screenshot_tmp = fname.startswith("screenshots_") and fname.endswith(".zip")
                        if is_export_tmp or is_screenshot_tmp:
                            fpath = os.path.join(export_dir, fname)
                            if now - os.path.getmtime(fpath) > 3600:
                                os.remove(fpath)
            except Exception:
                logger.warning("清理导出临时文件失败（忽略，下轮重试）", exc_info=True)

            # 兜底清理 openpyxl 泄漏的 /tmp 临时目录（超过 30 分钟未动）
            # 正常 wb.close() 会清掉；异常路径可能漏清，累积会吃满磁盘
            try:
                tmp_root = "/tmp"
                if os.path.isdir(tmp_root):
                    for fname in os.listdir(tmp_root):
                        if not fname.startswith("openpyxl."):
                            continue
                        fpath = os.path.join(tmp_root, fname)
                        try:
                            if now - os.path.getmtime(fpath) > 1800:
                                if os.path.isdir(fpath):
                                    import shutil as _sh
                                    _sh.rmtree(fpath, ignore_errors=True)
                                else:
                                    os.remove(fpath)
                                logger.info(f"清理 openpyxl 泄漏目录: {fname}")
                        except Exception:
                            logger.warning(f"清理 openpyxl 泄漏目录失败（忽略）: {fname}", exc_info=True)
            except Exception:
                logger.warning("扫描 /tmp openpyxl 泄漏目录失败（忽略，下轮重试）", exc_info=True)

            # 每 10 次循环（约 5 分钟）做一次 WAL checkpoint —— 按需 TRUNCATE 策略
            # 经 30k recon 实测：固定 TRUNCATE 会触发 ~200-400ms 的 commit 阻塞抖动，
            # 影响 worker pull_tasks 和 accept_results_batch（max hold 407ms）。
            # 新策略：
            #   默认走 PASSIVE（非阻塞，能 checkpoint 多少算多少）
            #   仅当 WAL 文件超过阈值才主动 TRUNCATE（兜底，几小时一次）
            # WAL 自身有 PRAGMA wal_autocheckpoint=1000 + journal_size_limit=64MB 兜底
            _wal_checkpoint_counter += 1
            if _wal_checkpoint_counter >= 10:
                _wal_checkpoint_counter = 0
                wal_size = 0
                try:
                    wal_path = config.DB_PATH + "-wal"
                    if os.path.exists(wal_path):
                        wal_size = os.path.getsize(wal_path)
                except Exception:
                    logger.warning("读取 WAL 文件大小失败（按 0 处理，走 PASSIVE）", exc_info=True)
                wal_mb = wal_size / 1024 / 1024
                # 阈值：WAL > 128MB 才主动 TRUNCATE；否则 PASSIVE 不阻塞 writer
                mode = "TRUNCATE" if wal_size > 128 * 1024 * 1024 else "PASSIVE"
                try:
                    res = await db.wal_checkpoint(mode)
                    if res:
                        logger.debug(
                            f"WAL checkpoint({mode}) wal={wal_mb:.1f}MB "
                            f"busy={res[0]} log={res[1]} checkpointed={res[2]}"
                        )
                except Exception as e:
                    logger.warning(f"WAL checkpoint 失败: {e}")
                # 分档告警
                if wal_size > 500 * 1024 * 1024:
                    logger.error(f"WAL 文件过大: {wal_mb:.0f}MB（>500MB）")
                elif wal_size > 200 * 1024 * 1024:
                    logger.warning(f"WAL 文件偏大: {wal_mb:.0f}MB（>200MB）")

            # 兜底扫描：把 DB 中到期可重试的 callback 重新入队
            # （服务重启、内存队列丢失、_completion_watcher 漏检测 等场景的安全网）
            try:
                if _callback_send_queue is not None:
                    now_str = now_ts()
                    due = await db.list_callback_due(now_str, limit=50)
                    for row in due:
                        try:
                            _callback_send_queue.put_nowait(row["id"])
                        except asyncio.QueueFull:
                            break  # 队列满了下次再说
            except Exception as e:
                logger.warning(f"callback 兜底扫描异常: {e}")

            # 兜底完成检测：扫描长期 running 的批次（防止入队事件丢失）
            # 只扫最近活动的 batches，避免全表查询
            try:
                # ORDER BY updated_at 单独一列不是全序：updated_at 是秒级精度，
                # 一次提交里的多个批次会拿到同一个时间戳，LIMIT 30 于是返回的是
                # **不确定的 30 行**（不只是顺序不定，是行集合不定）——SQLite 走
                # rowid、PG 走堆序，UPDATE 之后堆序还会漂。补 id DESC 把它变成全序。
                # NULLS LAST 是 SQLite DESC 的默认行为（实测），显式写出来只是为了
                # 让 PG 对齐（PG 的 DESC 默认 NULLS FIRST），SQLite 侧是 no-op。
                async with db.read() as rc, rc.execute(
                    """SELECT id FROM batches
                       WHERE status='running'
                       ORDER BY updated_at DESC NULLS LAST, id DESC
                       LIMIT 30"""
                ) as c:
                    rows = await c.fetchall()
                for r in rows:
                    _completion_check_set.add(r[0])
            except Exception as e:
                logger.warning(f"completion 兜底扫描异常: {e}")
        except Exception as e:
            logger.error(f"超时任务循环异常: {e}")


async def _completion_watcher():
    """完成检测协程：消费 _completion_check_set，找出已完成的批次。

    特点：
    - 完全脱离 worker 写入热路径
    - 每 _COMPLETION_WATCHER_INTERVAL 秒醒来一次，批量处理
    - 单次最多处理 100 个 batch_id，避免长时间占用事件循环
    - 任何异常只 log，不中断协程
    """
    logger.info("✅ 完成检测协程启动")
    while True:
        await asyncio.sleep(_COMPLETION_WATCHER_INTERVAL)
        try:
            # 一次性取走集合（用 set 自身的"原子读取"——单线程 asyncio 下安全）
            if not _completion_check_set:
                continue
            batch_ids = list(_completion_check_set)[:100]
            for bid in batch_ids:
                _completion_check_set.discard(bid)

            for bid in batch_ids:
                try:
                    # 快速 SQL：判断 task + screenshot 是否全部终态
                    snap = await db.get_batch_completion_status(bid)
                    if not snap["all_terminal"]:
                        continue
                    # 变体自动展开（仅 expand_variants 批次）：本轮全部终态后，把已采 ASIN
                    # 的同族变体入队【同一批次】继续采。新增 >0 则本批未真正完成，下轮再查。
                    added = await db.expand_batch_variants(bid)
                    if added > 0:
                        rnd = _expand_rounds.get(bid, 0) + 1
                        _expand_rounds[bid] = rnd
                        if rnd >= 2:
                            logger.warning(
                                f"🧬 批次 {bid} 变体展开第 {rnd} 轮仍新增 {added} 个任务："
                                f"家族列表可能不一致，请留意（去重保证仍会收敛）")
                        else:
                            logger.info(f"🧬 批次 {bid} 变体展开：新增 {added} 个同族任务，继续采集")
                        _completion_check_set.add(bid)  # 等这轮采完再复查
                        continue
                    _expand_rounds.pop(bid, None)
                    # 全部终态且无新变体 → 状态机转移（幂等）
                    changed = await db.mark_batch_completed(bid)
                    if changed:
                        logger.info(f"✅ 批次 {bid} 已完成，入队回调")
                        # 入队 callback dispatcher（仅当批次配置了回调）
                        if _callback_send_queue is not None:
                            try:
                                _callback_send_queue.put_nowait(bid)
                            except asyncio.QueueFull:
                                logger.warning("⚠️ callback 发送队列已满，依赖兜底扫描重发")
                except Exception as e:
                    logger.warning(f"完成检测单批次异常 batch_id={bid}: {e}")
        except Exception as e:
            logger.error(f"完成检测协程异常: {e}")


async def _callback_dispatcher():
    """回调发送协程：串行消费 _callback_send_queue + DB 待发条目，POST 到调用方。

    串行（而非并发）的原因：
    - 调用方接收端通常单实例，并发推送反而堆积
    - 串行简单可控，单次 10s 超时不会拖累整体
    - 失败按退避表延迟，写回 DB 等待下次扫描

    防卡死：
    - asyncio.wait_for(post, timeout=10) 强制中断慢响应
    - 重试上限 5 次后 callback_status='failed'，永不再处理
    """
    logger.info("✅ 回调发送协程启动")
    # 独立 httpx client（不与采集相关 HTTP 复用）
    async with httpx.AsyncClient(timeout=_CALLBACK_HTTP_TIMEOUT,
                                  follow_redirects=False) as client:
        while True:
            try:
                # 阻塞等下一个 batch_id（队列空时不空转）
                batch_id = await _callback_send_queue.get()
            except Exception as e:
                logger.error(f"回调队列读取异常: {e}")
                await asyncio.sleep(1)
                continue
            try:
                await _send_one_callback(client, batch_id)
            except Exception as e:
                logger.error(f"回调发送异常 batch_id={batch_id}: {e}")


async def _send_one_callback(client: httpx.AsyncClient, batch_id: int):
    """发送单个 batch 的回调。失败按退避表写回 DB。"""
    # 读取 batch 当前状态（要求 callback_status='pending' 才发；其他状态可能已 sent/failed/取消）
    async with db.read() as rc, rc.execute(
        """SELECT id, name, external_id, callback_url, callback_attempts,
                  callback_status, completed_at, status
           FROM batches WHERE id=?""",
        (batch_id,)
    ) as c:
        row = await c.fetchone()
    if not row:
        return
    batch = dict(row)
    if batch.get("callback_status") != "pending":
        return  # 已 sent / failed / disabled / 无回调
    callback_url = batch.get("callback_url")
    if not callback_url:
        return
    # 防御：完成检测协程可能比兜底扫描慢一拍，导致 callback_status='pending'
    # 但 status 还是 'running'。此时跳过，等下个周期。
    if batch.get("status") != "completed":
        return
    if not batch.get("completed_at"):
        return

    # 二次 SSRF 校验（防 DB 被改）
    ok, reason = await _is_safe_callback_url(callback_url)
    if not ok:
        await db.mark_callback_attempt(
            batch_id, success=False, error=f"unsafe_url:{reason}",
            max_attempts=1,  # 直接终态
        )
        return

    # 拉完成统计构造载荷
    snap = await db.get_batch_completion_status(batch_id)
    t = snap["tasks"]
    s = snap["screenshots"]
    total = t["total"] or 0
    success_rate = (t["done"] / total) if total > 0 else 0.0

    # 计算用时
    duration = None
    try:
        async with db.read() as rc, rc.execute(
            "SELECT created_at FROM batches WHERE id=?", (batch_id,)
        ) as c:
            row2 = await c.fetchone()
        if row2 and row2[0] and batch.get("completed_at"):
            start = datetime.strptime(row2[0][:19], '%Y-%m-%d %H:%M:%S')
            stop = datetime.strptime(batch["completed_at"][:19], '%Y-%m-%d %H:%M:%S')
            duration = int((stop - start).total_seconds())
    except Exception:
        logger.debug("计算批次耗时失败（duration 置空）", exc_info=True)

    server_base = (_runtime_settings.get("server_public_base") or "").rstrip("/")
    if not server_base:
        # 没配置公开 URL，就用 callback_url 的同 scheme（仅作 hint，不影响发送）
        server_base = ""
    data_url = f"{server_base}/api/results?batch_id={batch_id}" if server_base else None
    export_url = f"{server_base}/api/export/{batch['name']}" if server_base else None

    attempts = (batch.get("callback_attempts") or 0) + 1
    completed_at_str = batch.get("completed_at") or ""
    event_id = (
        f"evt_{batch_id}_"
        f"{completed_at_str.replace(' ','T').replace(':','').replace('-','')}"
    ) if completed_at_str else f"evt_{batch_id}"

    payload = {
        "event": "batch.completed",
        "batch_id": batch_id,
        "batch_name": batch["name"],
        "external_id": batch.get("external_id"),
        "status": batch.get("status") or "completed",
        "stats": {
            "total": total,
            "done": t["done"] or 0,
            "failed": t["failed"] or 0,
            "success_rate": round(success_rate, 4),
            "duration_seconds": duration,
        },
        "screenshots": {
            "total": s["total"] or 0,
            "done": s["done"] or 0,
            "failed": s["failed"] or 0,
        },
        "completed_at": batch.get("completed_at"),
        "data_url": data_url,
        "export_url": export_url,
    }
    headers = {
        "X-Scraper-Event-Id": event_id,
        "X-Scraper-Delivery-Attempt": str(attempts),
        "User-Agent": "amazon-scraper-v4/callback",
        "Content-Type": "application/json",
    }

    try:
        resp = await client.post(callback_url, json=payload, headers=headers)
        if 200 <= resp.status_code < 300:
            await db.mark_callback_attempt(batch_id, success=True,
                                            max_attempts=_CALLBACK_MAX_ATTEMPTS)
            logger.info(f"📬 callback 发送成功 batch={batch_id} attempt={attempts}")
            return
        err = f"HTTP {resp.status_code}: {(resp.text or '')[:200]}"
    except asyncio.TimeoutError:
        err = "timeout"
    except Exception as e:
        err = f"{type(e).__name__}: {e}"[:300]

    # 失败处理：计算下次重试时间
    delay_idx = min(attempts - 1, len(_CALLBACK_RETRY_DELAYS) - 1)
    delay = _CALLBACK_RETRY_DELAYS[delay_idx] if attempts < _CALLBACK_MAX_ATTEMPTS else 0
    next_retry = (
        ts_from(utc_now() + timedelta(seconds=delay))
        if attempts < _CALLBACK_MAX_ATTEMPTS else None
    )
    result = await db.mark_callback_attempt(
        batch_id, success=False, error=err,
        next_retry_at=next_retry, max_attempts=_CALLBACK_MAX_ATTEMPTS,
    )
    if result.get("final_status") == "failed":
        logger.warning(f"❌ callback 5 次失败终态 batch={batch_id} last_error={err}")
    else:
        logger.info(f"⏳ callback 失败 batch={batch_id} attempt={attempts}/{_CALLBACK_MAX_ATTEMPTS}，{delay}s 后重试")


async def _auto_scrape_scheduler():
    """定时自动采集调度（支持 interval_days 和文件指定 ASIN）"""
    while True:
        await asyncio.sleep(60)
        try:
            schedules = _runtime_settings.get("auto_scrape_schedules", [])
            now = _cn_now()  # 按中国时间判定「执行时间」，使 18:00 = 中国 18:00
            today = now.strftime("%Y-%m-%d")
            changed = False

            for sched in schedules:
                if not sched.get("enabled"):
                    continue
                sched_time = sched.get("time", "")
                if not sched_time:
                    continue
                try:
                    hour, minute = map(int, sched_time.split(":"))
                except ValueError:
                    continue

                # 检查时间点是否到达
                if now.hour < hour or (now.hour == hour and now.minute < minute):
                    continue

                # 检查间隔天数
                interval = sched.get("interval_days", 1)
                last_run = sched.get("last_run_date", "")
                if last_run:
                    try:
                        last_date = datetime.strptime(last_run, "%Y-%m-%d")
                        if (now - last_date).days < interval:
                            continue
                    except ValueError:
                        pass

                # 获取 ASIN 列表：优先从文件，否则全库
                source_file = sched.get("source_file", "")
                if source_file and os.path.isfile(source_file):
                    asins = _extract_asins_from_file(source_file)
                else:
                    asins = await db.get_all_asins()

                if not asins:
                    continue

                sched_name = sched.get("name", "task")
                # P4.7：**分钟 -> 秒**（有意的行为改动，见 _batch_name 的 docstring）。
                # 分钟精度下「同一分钟内自动 + 手动触发同一个定时任务」会撞名，
                # 而撞名不是 no-op：create_batch 返回既有批次 id，新 ASIN 被
                # 悄悄塞进上一个批次。
                batch_name = _batch_name(f"auto_{sched_name}")
                zc = _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE)
                ns = sched.get("needs_screenshot", False)
                batch_id = await db.create_batch(batch_name, ns, is_auto=True)
                await db.create_tasks(batch_id, asins, zc, ns)
                sched["last_run_date"] = today
                changed = True
                logger.info(f"自动采集已调度: {batch_name}, {len(asins)} ASINs (间隔{interval}天)")

            if changed:
                _save_settings()
        except Exception as e:
            logger.error(f"自动采集调度异常: {e}")


# ==================== 全局并发协调 ====================

def _allocate_quotas():
    """根据 worker 健康度分配并发配额"""
    active = {wid: info for wid, info in _global_coordinator.items()
              if wid in _worker_registry}
    if not active:
        return

    max_conc = _runtime_settings.get("global_max_concurrency", config.GLOBAL_MAX_CONCURRENCY)
    max_qps = _runtime_settings.get("global_max_qps", config.GLOBAL_MAX_QPS)

    # 健康度加权分配
    scores = {}
    for wid, info in active.items():
        metrics = info.get("metrics", {})
        sr = metrics.get("success_rate", 1.0)
        br = metrics.get("block_rate", 0.0)
        scores[wid] = max(0.1, sr * (1 - br * 5))

    total_score = sum(scores.values())
    per_worker_max_conc = _runtime_settings.get("max_concurrency", config.MAX_CONCURRENCY)
    per_worker_max_qps = _runtime_settings.get("token_bucket_rate", config.TOKEN_BUCKET_RATE)

    for wid in active:
        weight = scores[wid] / total_score if total_score > 0 else 1.0 / len(active)
        allocated_conc = max(config.MIN_CONCURRENCY, int(max_conc * weight))
        allocated_qps = max(1.0, max_qps * weight)
        active[wid]["quota"] = {
            # 取 global 分配值和单 worker 上限的较小值
            "max_concurrency": min(allocated_conc, per_worker_max_conc),
            "max_qps": min(allocated_qps, per_worker_max_qps),
        }


# ==================== HTML 页面 ====================
#
# 5 个页面搬到 server/api/pages.py（Phase 3.1）。router 光秃 —— 不带 tags/prefix，
# 因为 /openapi.json 是黄金基线的一步、逐字节钉死，加 tags 会让 51 个 path 一起飘红。
# `templates` 与 `| cst` 过滤器一并搬走（全仓只有那 5 个 handler 用）；
# `app.mount("/static", ...)` 是 app 级挂载不是路由，留在上面原地。
from server.api import pages as _pages_api  # noqa: E402

app.include_router(_pages_api.router)


# ==================== API: 批次和任务 ====================

MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50MB 上限，防止 2GB 内存 VPS OOM

#: ``POST /api/upload`` 撞名时的**机器读**错误码，放在 409 响应体的
#: ``detail.error`` 里。人读的说明在 ``detail.message``。
#:
#: 这个码走的是 ``HTTPException``，不经过 ``server/api/sync.py`` 的 ``_err``，
#: 所以 ``tests/test_error_codes.py`` 那道 AST 扫描**扫不到它** —— 与全局 500 的
#: ``"internal_error"`` 完全同例。两者都由该文件里一条**显式**的注册断言看守
#: （``test_batch_name_conflict_is_registered`` / ``test_internal_error_is_registered``）。
#: 改这里的字面量、或把它从 ``ERROR_CODES`` 里删掉，那条用例当场红。
_BATCH_NAME_CONFLICT_CODE = "batch_name_conflict"


# ==================== 完成通知 / Callback 基础设施 ====================

# 内存中需要"检查是否完成"的 batch_id 集合：
# - 由 accept_results_batch 在写入结果后 add(batch_id) （O(1)、不阻塞）
# - 由 _completion_watcher 协程消费
_completion_check_set: Set[int] = set()
# 变体自动展开：记录每个批次已展开的轮次。正常 2 轮收敛（轮1=上传，轮2=变体）；
# 若第 2 次展开仍新增（=采集第 3 轮）则打 WARNING（家族列表不一致信号），但靠
# UNIQUE(batch_id,asin) 去重仍保证收敛、不死循环。批次完成后清理。
_expand_rounds: Dict[int, int] = {}
# 待发送的 callback 内存队列：
# - _completion_watcher 把刚标记 completed 的 batch_id 入队
# - _timeout_task_loop 兜底从 DB 扫描入队（处理重启 / 实时漏掉的）
# - _callback_dispatcher 串行消费 + HTTP 发送
_callback_send_queue: "asyncio.Queue[int]" = None  # 在 lifespan 启动时实例化

# 回调重试退避（秒）
_CALLBACK_RETRY_DELAYS = [30, 300, 1800, 7200]   # 第 1/2/3/4 次失败后的等待
_CALLBACK_MAX_ATTEMPTS = 5
_CALLBACK_HTTP_TIMEOUT = 10                       # 单次 HTTP 超时
_COMPLETION_WATCHER_INTERVAL = 2.0                # 每 2 秒批量处理一次

# SSRF 防护：拒绝指向内网的 URL
_SSRF_BLOCKED_HOSTS = {"localhost", "ip6-localhost", "ip6-loopback"}


def _ip_is_blocked(ip: "ipaddress._BaseAddress") -> bool:
    """内网 / 回环 / 链路本地 / 保留 / 组播 地址一律拒绝。"""
    return (ip.is_loopback or ip.is_private or ip.is_link_local
            or ip.is_reserved or ip.is_multicast or ip.is_unspecified)


async def _is_safe_callback_url(url: str) -> tuple[bool, str]:
    """判定 callback URL 是否安全可用于公网回调。

    返回 (ok, reason)。公网场景下：
    - 只允许 http/https
    - 拒绝 localhost 字面量
    - 拒绝字面量私网 / 回环 / 链路本地 IP（10/8、172.16-31、192.168/16、127/8、169.254/16、::1 等）
    - 域名做 DNS 解析后复检：任一解析结果落在内网 / 元数据 IP（如 169.254.169.254）
      即拒绝，堵住"域名指向内网"的 SSRF 绕过。
      注：解析与实际发起连接之间仍存在 TOCTOU 窗口（DNS 可能被改），这里只做尽力校验；
      真正的强隔离需在连接层 pin 住已校验 IP。
    """
    if not url or not isinstance(url, str):
        return False, "empty"
    try:
        u = urlparse(url.strip())
    except Exception as e:
        return False, f"invalid_url:{e}"
    if u.scheme not in ("http", "https"):
        return False, f"bad_scheme:{u.scheme}"
    host = (u.hostname or "").lower()
    if not host:
        return False, "missing_host"
    if host in _SSRF_BLOCKED_HOSTS:
        return False, "blocked_host"
    # 尝试作为 IP 字面量解析
    try:
        ip = ipaddress.ip_address(host)
        if _ip_is_blocked(ip):
            return False, f"blocked_ip:{host}"
        return True, ""  # 合法公网 IP 字面量，无需 DNS 解析
    except ValueError:
        pass  # 是域名，继续做 DNS 复检

    # 域名：解析所有 A/AAAA 记录，任一落在内网即拒绝
    try:
        port = u.port or (443 if u.scheme == "https" else 80)
    except ValueError:
        return False, "bad_port"
    try:
        loop = asyncio.get_running_loop()
        infos = await loop.getaddrinfo(host, port, type=socket.SOCK_STREAM)
    except Exception as e:
        return False, f"dns_fail:{type(e).__name__}"
    if not infos:
        return False, "dns_empty"
    for info in infos:
        ip_str = info[4][0]
        try:
            ip = ipaddress.ip_address(ip_str)
        except ValueError:
            continue
        if _ip_is_blocked(ip):
            return False, f"blocked_resolved_ip:{ip_str}"
    return True, ""


# `_ASIN_RE` 现在从 common/core/idents.py import（见文件头），本地不再定义。
# 那份真源与 worker/parser.py 共用；`.strip().upper()` 的归一留在下面
# `_normalize_asin` 里 —— 正则自己不做大小写归一。
# 美国邮编：5 位数字（兼容 ZIP+4，前 5 位）
#
# P4.6：``\Z`` 而不是 ``$``。Python 的 ``$`` 在**末尾恰好一个换行**处也匹配，
# 于是 ``'10001\n'`` 能通过一条看起来是「只接受 5 位数字」的校验。
# 本函数在这里之前已经 ``.strip()`` 过，所以这一改**对任何输入都不改变结果**
# （已实测），它守的是「将来有人把 strip 挪走 / 复制这条正则去别处用」。
_US_ZIP_RE = re.compile(r'^\d{5}\Z')


def _normalize_asin(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip().upper()
    return s if _ASIN_RE.match(s) else None


def _normalize_zip(val) -> Optional[str]:
    """规范化邮编。返回 5 位数字字符串，无效则 None。
    支持去掉前导/尾随空白、ZIP+4（取前 5 位）、Excel 数字单元格（如 10001 不会带前导 0 的损失）。
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Excel 数字 90001.0 → "90001"
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    # 前 5 位（兼容 "10001-1234"）
    head = s.split("-", 1)[0].strip()
    # 数字邮编位数补 0（Excel 把 "01234" 存为 1234）
    # P4.6：补零规则的唯一真源是 common/core/zipcode.py —— 本函数归一出来的值
    # 会去和 relay 归一出来的 zip_requested 比对，两边一边补零一边不补，
    # 就是凭空造出的 mismatch。原先这里写的是 ``len(head) <= 5``，
    # 而 ``"12345".zfill(5) == "12345"``，所以 ``or head`` 这一支覆盖的
    # 「5 位数字」情形结果完全一致（已实测）。
    head = _zfill_short_numeric(head) or head
    return head if _US_ZIP_RE.match(head) else None


def _safe_fs_component(name: str) -> Optional[str]:
    """校验一个字符串是否可安全用作文件系统路径组件（目录名 / 文件名前缀）。

    截图存储路径由 batch_name / asin 直接拼成（server/static/screenshots/<batch>/<asin>.png）。
    含 '/'、'\\'、'..'、控制字符或空值的输入会造成路径穿越，一律拒绝（返回 None）。
    合法的自动/手动批次名（字母数字、下划线、连字符、点、空格、中文）不受影响。
    """
    if not name or not isinstance(name, str):
        return None
    if "/" in name or "\\" in name or "\x00" in name:
        return None
    # 拒绝 '.'、'..' 以及任何以 '..' 开头的穿越尝试
    if name in (".", "..") or ".." in name:
        return None
    return name


@app.post("/api/upload")
async def api_upload(request: Request,
                     file: UploadFile = File(...),
                     batch_name: str = Form(None),
                     zip_code: str = Form(None),
                     needs_screenshot: bool = Form(False),
                     callback_url: str = Form(None),
                     external_id: str = Form(None),
                     expand_variants: bool = Form(False)):
    """上传 ASIN 文件创建批次。

    支持 xlsx / csv / txt 三种格式：
    - **xlsx / csv**：A 列 = ASIN，B 列（可选）= 该 ASIN 单独的采集邮编（5 位数字）
      B 列为空 → 用本次上传指定的 batch zip（或服务端默认）
    - **txt**：每行一个 ASIN（无邮编列）

    可选 callback：
    - `callback_url`：批次完成时（含截图）POST 到此 URL 通知调用方
    - `external_id`：调用方自己的批次 ID，原样回传，便于追踪

    批次名冲突 → **409 Conflict**，绝不静默合并进已有批次。409 的响应体里带
    `detail.batch_id` / `detail.batch_name` / `detail.status_url`，调用方可以
    直接拿去接着轮询。

    由此得到的三条性质（调用方可以依赖）：
    - **本端点可以安全重试**：网络超时后重发，若上一次其实成功了，拿到的是
      409 + 那个既有 batch_id，而不是又建一个批次、也不会悄悄并进去。
    - 批次名不需要毫秒精度去躲开合并。
    - **200 恒等于「新建了一个批次」**，`inserted` 因此不再有歧义。

    200 响应里的 `external_id` / `callback_url` 回显的是**批次实际存下来的**值
    （读回后再回显），不是请求里的值。
    """
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"文件过大：{len(content)//1024//1024}MB，上限 {MAX_UPLOAD_BYTES//1024//1024}MB")
    filename = file.filename or ""

    asins: List[str] = []                # 顺序、可重复（去重在后面）
    per_asin_zip: Dict[str, str] = {}    # 仅记录 B 列指定的邮编
    invalid_zip_count = 0

    def add_pair(asin_val, zip_val):
        nonlocal invalid_zip_count
        asin = _normalize_asin(asin_val)
        if not asin:
            return
        asins.append(asin)
        if zip_val is not None and str(zip_val).strip():
            zip_norm = _normalize_zip(zip_val)
            if zip_norm:
                # 同一 ASIN 在多行重复时，以第一次出现的非空 zip 为准
                per_asin_zip.setdefault(asin, zip_norm)
            else:
                invalid_zip_count += 1

    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row:
                    continue
                a = row[0] if len(row) > 0 else None
                b = row[1] if len(row) > 1 else None
                # 兼容旧上传：当 A 列不是 ASIN 时，扫描该行所有列找 ASIN（不带 zip）
                if _normalize_asin(a):
                    add_pair(a, b)
                else:
                    for cell in row:
                        if _normalize_asin(cell):
                            add_pair(cell, None)
        finally:
            wb.close()
    elif filename.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if not row:
                continue
            a = row[0] if len(row) > 0 else None
            b = row[1] if len(row) > 1 else None
            if _normalize_asin(a):
                add_pair(a, b)
            else:
                for cell in row:
                    if _normalize_asin(cell):
                        add_pair(cell, None)
    else:
        text = content.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            add_pair(line, None)

    if not asins:
        raise HTTPException(400, "未找到有效 ASIN")

    # 去重（保持顺序）
    seen = set()
    unique_asins = []
    for a in asins:
        if a not in seen:
            unique_asins.append(a)
            seen.add(a)

    if not batch_name:
        batch_name = _batch_name("batch")     # P4.7：精度不变（本来就是秒）

    # callback_url 校验（防 SSRF + 格式）
    cb_url = (callback_url or "").strip() or None
    if cb_url:
        ok, reason = await _is_safe_callback_url(cb_url)
        if not ok:
            raise HTTPException(400, f"非法 callback_url（{reason}）。仅接受 http(s)://公网域名/IP")

    ext_id = (external_id or "").strip()[:120] or None  # 上限 120 字符防滥用

    zc = zip_code or _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE)

    # 构造 status URL（调用方可以轮询）。409 也要带上它，所以在建批次之前先算好。
    base = str(request.base_url).rstrip("/")
    status_url = f"{base}/api/batches/{batch_name}/status"

    batch_id, created = await db.create_batch_if_absent(
        batch_name, needs_screenshot,
        external_id=ext_id, callback_url=cb_url,
        expand_variants=expand_variants,
    )
    if not created:
        # 撞名 -> 409，**绝不静默合并**。
        #
        # 以前这里是 200 + 既有 batch_id，实测后果（Phase 4.7 已证明撞名不是
        # no-op，本轮修的就是它）：
        #   * 本次的新 ASIN 被悄悄塞进上一个批次，「一次采集」的语义破掉；
        #   * `inserted` 在部分重叠时是非零，看起来像成功新建；
        #   * **本次的 external_id / callback_url 被静默丢弃**（INSERT OR IGNORE
        #     整行不插，既有行一个字段都不更新），而响应回显的是**请求**里的值 ——
        #     调用方以为回调注册好了，回调永远不会触发。
        #
        # 为什么是 409，而不是「自动加后缀」或「200 加个 merged 标志」：只有让
        # 撞名变成一个**可识别的失败**，POST /api/upload 才是可安全重试的 ——
        # 网络超时后重发，上一次若其实成功了，这里回的是 409 + 那个 batch_id。
        # 自动加后缀会让重试造出第二个批次，200+标志则要求每个调用方都记得读那个
        # 标志，漏读的代价与今天一模一样。
        #
        # 注意 batch_id 是**既有批次**的 id（create_batch_if_absent 撞名时照样把它
        # SELECT 回来），调用方可以直接拿 status_url 接着轮询，不必再查一次。
        raise HTTPException(409, {
            "error": _BATCH_NAME_CONFLICT_CODE,
            "message": f"批次名已存在: {batch_name}（未合并，也未改动既有批次）",
            "batch_id": batch_id,
            "batch_name": batch_name,
            "status_url": status_url,
        })

    inserted = await db.create_tasks(
        batch_id, unique_asins, zc, needs_screenshot,
        per_asin_zip=per_asin_zip,
    )

    # 回显**存下来的**值，不是请求里的值。
    # 撞名已经在上面 409 掉了，所以走到这里两者必然相等——但「回显请求值」这个
    # 写法本身就是上面那个回调撒谎 bug 的载体，留着它等于把地雷埋回去。
    # 读回一行的代价（一次主键级 SELECT）远小于「回调注册成功了吗」这种问题。
    stored = await db.get_batch_by_name(batch_name)
    return {
        "batch_id": batch_id,
        "batch_name": batch_name,
        "external_id": stored.get("external_id") if stored else ext_id,
        "total_asins": len(unique_asins),
        "inserted": inserted,
        "per_asin_zip_count": len(per_asin_zip),
        "invalid_zip_rows": invalid_zip_count,
        "callback_url": stored.get("callback_url") if stored else cb_url,
        "status_url": status_url,
    }


# ==================== F-009: 卖家店铺采集 ====================
#
# 4 个端点搬到 server/api/sellers.py（Phase 3.4）：upload-sellers、
# seller-batches/{id}/progress、seller-batches/{id}/discoveries，
# 外加 POST /api/tasks/seller-result —— 它原本待在下面的「Worker 任务拉取和
# 提交」节里，节头骗人，按域它属于 F-009，所以一起搬走。
#
# ⚠ 归域是逐端点判定的：紧跟本节头的 @app.get("/api/batches") 往下 300+ 行
#   是被挤下来的批次 / worker / 设置端点，**不属于**卖家采集，留在原地。
#
# 这一族黄金 78 步一步都没覆盖（upload-sellers 的响应含逐次不同的批次名，
# 补不进基线），替代网是 tests/test_seller_api.py，两个后端都跑。
from server.api import sellers as _sellers_api  # noqa: E402

app.include_router(_sellers_api.router)


@app.get("/api/batches")
async def api_batches():
    batches = await db.get_batches()
    return {"batches": batches}


@app.get("/api/progress")
async def api_progress(batch_id: int = None):
    return await db.get_progress(batch_id)


@app.get("/api/batches/{batch_name}/screenshots/progress")
async def api_screenshot_progress(batch_name: str):
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    return await db.get_screenshot_progress(batch["id"])


@app.get("/api/batches/{batch_name}/status")
async def api_batch_status(batch_name: str):
    """轮询批次完成状态（兼容旧调用方 + 给纯轮询模式用）。

    返回结构：
        {
          "batch_name": ...,
          "batch_id": ...,
          "external_id": ...,
          "status": "running"|"completed"|"failed",
          "stats": {total, done, failed, success_rate, duration_seconds},
          "screenshots": {total, done, failed},
          "completed_at": null|"2026-...",
          "callback": {url, status, attempts, last_error, sent_at}
        }
    """
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    snap = await db.get_batch_completion_status(batch["id"])
    t = snap["tasks"]
    s = snap["screenshots"]
    total = t["total"] or 0
    success_rate = (t["done"] / total) if total > 0 else 0.0

    # 持续时长（创建到完成；未完成则到当前时间）
    duration = None
    try:
        created_at = batch.get("created_at")
        end_at = batch.get("completed_at") or now_ts()
        if created_at:
            start = datetime.strptime(created_at[:19], '%Y-%m-%d %H:%M:%S')
            stop = datetime.strptime(end_at[:19], '%Y-%m-%d %H:%M:%S')
            duration = int((stop - start).total_seconds())
    except Exception:
        logger.debug("计算批次耗时失败（duration 置空）", exc_info=True)

    return {
        "batch_name": batch_name,
        "batch_id": batch["id"],
        "external_id": batch.get("external_id"),
        "status": batch.get("status") or "running",
        "stats": {
            "total": total,
            "done": t["done"] or 0,
            "failed": t["failed"] or 0,
            "open": t["open"] or 0,
            "success_rate": round(success_rate, 4),
            "duration_seconds": duration,
        },
        "screenshots": {
            "total": s["total"] or 0,
            "done": s["done"] or 0,
            "failed": s["failed"] or 0,
            "open": s["open"] or 0,
        },
        "completed_at": batch.get("completed_at"),
        "callback": {
            "url": batch.get("callback_url"),
            "status": batch.get("callback_status"),
            "attempts": batch.get("callback_attempts") or 0,
            "last_error": batch.get("callback_last_error"),
            "sent_at": batch.get("callback_sent_at"),
            "next_retry_at": batch.get("callback_next_retry_at"),
        },
    }


@app.post("/api/batches/{batch_name}/callback/retry")
async def api_batch_callback_retry(batch_name: str):
    """运维手动触发：重置该批次的 callback 状态，立即重新发送。
    可用于：webhook 失败 5 次后；或调用方端点恢复后想重新接收一次通知。
    """
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    if not batch.get("callback_url"):
        raise HTTPException(400, "该批次没有配置 callback_url")
    changed = await db.reset_callback_for_retry(batch["id"])
    # 入队让 dispatcher 立刻处理
    if _callback_send_queue is not None:
        try:
            _callback_send_queue.put_nowait(batch["id"])
        except Exception:
            logger.warning(f"callback 重发入队失败（忽略）: batch_id={batch['id']}", exc_info=True)
    return {"ok": changed, "batch_id": batch["id"]}


@app.post("/api/batches/{batch_id}/prioritize")
async def api_prioritize(batch_id: int):
    await db.prioritize_batch(batch_id)
    return {"ok": True}


@app.post("/api/batches/{batch_name}/retry")
async def api_retry_batch(batch_name: str, force: bool = False):
    """重试失败任务

    始终跳过 NO_AUTO_RETRY_ERROR_TYPES（如 variant_offset），因为这些类型
    是 Amazon 侧返回兄弟变体页的稳定问题，不再重试。
    返回 retried/skipped_no_retry 数量，前端可展示。
    """
    from common.core import NO_AUTO_RETRY_ERROR_TYPES
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    batch_id = batch["id"]

    # 排除清单在**这里**算好再传下去：它是本端点的策略（"始终跳过
    # variant_offset"），而且要原样回显在响应体里。db 层只负责按清单执行。
    no_retry_list = sorted(NO_AUTO_RETRY_ERROR_TYPES)

    # force 参数保留兼容旧调用，但不覆盖 NO_AUTO_RETRY_ERROR_TYPES。
    res = await db.retry_failed_tasks(batch_id, no_retry_list)
    return {
        "ok": True,
        "retried": res["retried"],
        "skipped_no_retry": res["skipped"],
        "no_retry_types": no_retry_list,
        "forced": force,
    }


@app.delete("/api/batches/{batch_name}")
async def api_delete_batch(batch_name: str):
    """删除批次及其任务"""
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    batch_id = batch["id"]
    screenshot_files = await db.delete_batches([batch_id])

    # 删除物理截图文件
    _remove_screenshot_files(screenshot_files)
    return {"ok": True}


@app.post("/api/batches/delete-bulk")
async def api_delete_batches_bulk(request: Request):
    """批量删除多个批次（按 batch_id）及其全部关联数据 + 截图文件。
    入参 JSON：{"batch_ids": [1,2,3]}。一次事务删除，原子性。"""
    body = await request.json()
    raw = body.get("batch_ids", [])
    if not isinstance(raw, list):
        raise HTTPException(400, "batch_ids 必须是数组")
    # 仅接受整数 id，去重，上限保护（防超大 IN 子句）
    seen = set()
    batch_ids = []
    for x in raw:
        try:
            i = int(x)
        except (ValueError, TypeError):
            continue
        if i not in seen:
            seen.add(i)
            batch_ids.append(i)
    batch_ids = batch_ids[:500]
    if not batch_ids:
        raise HTTPException(400, "batch_ids 为空或无效")

    screenshot_files = await db.delete_batches(batch_ids)

    _remove_screenshot_files(screenshot_files)
    logger.info(f"批量删除批次: {len(batch_ids)} 个 (ids={batch_ids[:20]}{'...' if len(batch_ids) > 20 else ''})")
    return {"ok": True, "deleted": len(batch_ids)}


# ⚠ 下面这个端点是**旧接口**，新接入方一律用 /api/batches/{batch_id}/failures。
#
# 指向 /failures 这件事只写进 docstring（= /openapi.json 的 description），
# **没有**往响应体里加 deprecation 提示字段。理由是实测出来的，不是口味：
#   * `errors_batch_a` 是黄金 78 步里的一步（body 录着
#     {"error_summary": [], "failed_tasks": []}），加字段当场改基线，
#     而这一步与本轮要改的撞名行为毫无关系；
#   * 更要紧的是它是**线上格式变更**：今天在跑的调用方拿到一个没见过的键，
#     严格反序列化的那种会直接崩。为了一句提示去冒这个险不划算。
# docstring 走的是 /openapi.json（那**也**是黄金的一步），但它只动
# description 字符串、不动任何响应形状，对在跑的调用方是零风险。
# 「docstring 里必须指得出 /failures」由
# tests/test_errors_endpoint_points_at_failures.py 钉住。
@app.get("/api/batches/{batch_name}/errors")
async def api_batch_errors(batch_name: str):
    """获取批次错误详情（**旧接口，请改用 `/api/batches/{batch_id}/failures`**）。

    本端点有两条硬限制，都是设计如此、调大参数也没用：

    - `failed_tasks` **最多 200 条**，超出的部分直接看不见；
    - 排序是 `updated_at DESC, id DESC` —— 一次批量提交里的失败任务共享同一个
      秒级 `updated_at`，所以「最近 200 条」在同一批内部并不是一个有业务含义的切片。

    需要**完整**失败明细请用 `GET /api/batches/{batch_id}/failures`：按 batch_id
    取（不依赖批次名）、`limit` 上限 100000、不截断到 200 条，还支持
    `error_type` 过滤。

    保留本端点只是为了不打断已在使用它的调用方，不再接受新接入。
    """
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    batch_id = batch["id"]
    async with db.read() as rc:
        # cnt 并列时顺序原本不定 → 用 error_type 兜底成全序。
        # NULLS FIRST 是 SQLite ASC 的默认行为（实测），PG 的 ASC 默认 NULLS LAST，
        # 所以必须显式写；error_type 在两边都是二进制序（PG 侧 COLLATE "C"）。
        async with rc.execute(
            "SELECT error_type, COUNT(*) as cnt FROM tasks "
            "WHERE batch_id=? AND status='failed' "
            "GROUP BY error_type ORDER BY cnt DESC, error_type NULLS FIRST",
            (batch_id,)
        ) as c:
            error_summary = [dict(r) for r in await c.fetchall()]
        # 这一条原来是 ORDER BY updated_at DESC LIMIT 200，没有 tiebreaker。
        # updated_at 是秒级精度，而 accept_results_batch 会把**整次提交**盖上同一个
        # 时间戳：一批 260 个任务一起失败时，200 行的**行集合**在两个后端不一样
        # （实测 60 行不同），同一个后端换一次表重组也可能变。补 id DESC 后与
        # /api/batches/{id}/failures（get_batch_failures，本来就是这个写法）一致。
        async with rc.execute(
            "SELECT asin, error_type, error_detail, retry_count, worker_id, updated_at "
            "FROM tasks WHERE batch_id=? AND status='failed' "
            "ORDER BY updated_at DESC NULLS LAST, id DESC LIMIT 200",
            (batch_id,)
        ) as c:
            failed_tasks = [dict(r) for r in await c.fetchall()]
    return {"error_summary": error_summary, "failed_tasks": failed_tasks}


@app.get("/api/batches/{batch_id}/failures")
async def api_batch_failures(
    batch_id: int,
    error_type: Optional[str] = Query(None, description="逗号分隔的 error_type 过滤"),
    limit: int = Query(100000, ge=1, le=100000),
):
    """按 batch_id 获取失败任务明细；不依赖批次名，且不截断到 200 条。"""
    error_types = None
    if error_type:
        error_types = [t.strip() for t in error_type.split(",") if t.strip()]
    failed_tasks = await db.get_batch_failures(batch_id, error_types=error_types, limit=limit)
    return {"batch_id": batch_id, "failed_tasks": failed_tasks, "count": len(failed_tasks)}


# ==================== API: Worker 机群（注册表 / 心跳 / 配额）====================
#
# 6 个端点搬到 server/api/fleet.py（Phase 3.3）：coordinator、清离线、restart、
# worker/sync、workers、delete worker。router 光秃，不带 tags/prefix。
# _register_worker / _allocate_quotas 仍留在本文件（它们直接读写下面那几个模块级
# 全局，而黄金夹具按名字给 server.app 打补丁），fleet.py 走 _srv() 调它们。
# 路径全无遮蔽：/api/workers 是静态的，与 /api/workers/{worker_id} 不互吃。
from server.api import fleet as _fleet_api  # noqa: E402

app.include_router(_fleet_api.router)


@app.post("/api/settings/reset")
async def api_reset_settings():
    """恢复默认设置"""
    global _runtime_settings, _settings_version
    _runtime_settings = _default_settings()
    _settings_version += 1
    _save_settings()
    return {"ok": True, "settings": _runtime_settings}


# ==================== API: Worker 任务拉取和提交 ====================
#
# 6 个端点搬到 server/api/worker_queue.py（Phase 3.5）：tasks/pull、tasks/release、
# tasks/result、tasks/result/batch、tasks/screenshot、tasks/screenshot/fail。
# router 光秃，不带 tags/prefix。db / _worker_registry / _completion_check_set
# 仍留在本文件（黄金夹具与 PG 夹具都按名字给 server.app 打补丁），
# worker_queue.py 一律走 _srv()；_register_worker / _rollback_quietly /
# _normalize_asin / _safe_fs_component 同理留在本文件。
# 六条全是静态路径，/api/tasks 下没有 catch-all，注册次序不影响匹配。
from server.api import worker_queue as _worker_queue_api  # noqa: E402

app.include_router(_worker_queue_api.router)


# ==================== API: Worker 同步 ====================
# /api/worker/sync、GET /api/workers、DELETE /api/workers/{worker_id} 已搬到
# server/api/fleet.py（Phase 3.3）；router 在上面「Worker 机群」那一节挂载。


# ==================== API: 结果查询 / 结果删除 ====================
#
# 5 个端点搬到 server/api/results.py（Phase 3.6）：GET /api/results、
# GET /api/results/{asin}、GET /api/changes/stats，外加两条删除
# —— POST /api/results/delete-by-file 与 DELETE /api/results，
# 它们原本待在下面「诊断 / 侦查」节头之后，节头骗人，按域属于结果面。
# router 光秃，不带 tags/prefix。db / MAX_UPLOAD_BYTES / _remove_screenshot_files
# 仍留在本文件，results.py 一律走 _srv()。
# Phase 3.8 批 (3) 之后 results.py 里一条 SQL 都没有了：三段裸 SQL 收进了
# db.find_asins_by_search / db.get_batch_asin_set / db.delete_asins。
# 那段自拼的 f-string LIKE 曾经靠 common/pgdb/pool.py 的 _LIKE_QMARK_RE 按字面
# 文本改写撑着 PG 语义（D-16）；现在删除路径与读路径引用同一个谓词常量，
# 守卫是 tests/test_search_like_escape_parity.py（读/删选中同一批行，跑两列）。
# 路由匹配不变：GET /api/results/{asin} 与 POST /api/results/delete-by-file
# 方法不同，GET/DELETE /api/results 同理，两者相对顺序也与拆分前一致。
from server.api import results as _results_api  # noqa: E402

app.include_router(_results_api.router)


# ==================== API: 设置 ====================

@app.get("/api/settings")
async def api_get_settings():
    return {"settings": _runtime_settings, "version": _settings_version}


@app.put("/api/settings")
async def api_update_settings(request: Request):
    global _settings_version
    body = await request.json()

    for key, value in body.items():
        if key in _runtime_settings:
            _runtime_settings[key] = value

    _settings_version += 1
    _save_settings()
    return {"ok": True, "version": _settings_version, "settings": _runtime_settings}


# ==================== API: 导出 ====================
#
# 4 个端点搬到 server/api/export.py（Phase 3.7）：export/fields、export/all、
# export/{batch_name}（catch-all）、export/{batch_name}/screenshots，
# 连同 _parse_selected_fields / _get_export_headers / _prepare_row /
# _export_xlsx_streaming / _export_csv_streaming 等私有助手。
# router 光秃，不带 tags/prefix。db / logger / _cn_now / _safe_fs_component
# 仍留在本文件，export.py 一律走 _srv()。
#
# ⚠ /api/export/incremental 的注册顺序**不再由这里保证**（顺序局部化）：
#   它现在由 export.py 在自己第一个 @router.get 之前 include 进来，
#   所以「增量端点排在 catch-all 之前」这条不变量由单文件自上而下阅读保证，
#   下面这份 include 列表怎么重排都打不破它。守卫见
#   tests/test_incremental_export.py::RouteOrderTests（结构 + 行为 + 源码三层）
#   与 tools/phase5_preflight.py:check_route_order。
from server.api import export as _export_api  # noqa: E402

app.include_router(_export_api.router)


# ==================== API: 诊断 / 侦查 ====================
#
# 5 个端点（/api/diagnostic、/api/_debug/* 三个、DELETE /api/database）
# 搬到 server/api/debug.py（Phase 3.2）。router 光秃，不带 tags/prefix、
# 不在 router 上设 include_in_schema —— event-stream 那条的
# include_in_schema=False 是装饰器级参数，跟着函数走。
# 这一族路径全是静态的，`/api` 下没有一级 catch-all，注册次序不影响匹配。
from server.api import debug as _debug_api  # noqa: E402

app.include_router(_debug_api.router)


# ==================== 定时采集管理 ====================

_SCHEDULES_DIR = os.path.join(config.PROJECT_DIR, "data", "schedules")


def _get_schedules() -> list:
    return _runtime_settings.get("auto_scrape_schedules", [])


def _save_schedules(schedules: list):
    _runtime_settings["auto_scrape_schedules"] = schedules
    _save_settings()


def _extract_asins_from_file(filepath: str) -> list:
    """从文件提取 ASIN 列表"""
    asins = []
    seen = set()
    if not os.path.isfile(filepath):
        return asins
    with open(filepath, "rb") as f:
        content = f.read()
    filename = filepath.lower()
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                for cell in row:
                    if cell:
                        val = str(cell).strip().upper()
                        if re.match(r'^B[0-9A-Z]{9}$', val) and val not in seen:
                            asins.append(val)
                            seen.add(val)
        finally:
            wb.close()
    elif filename.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            for cell in row:
                val = cell.strip().upper()
                if re.match(r'^B[0-9A-Z]{9}$', val) and val not in seen:
                    asins.append(val)
                    seen.add(val)
    else:
        text = content.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            val = line.strip().upper()
            if re.match(r'^B[0-9A-Z]{9}$', val) and val not in seen:
                asins.append(val)
                seen.add(val)
    return asins


@app.get("/api/schedules")
async def api_list_schedules():
    return {"schedules": _get_schedules()}


@app.post("/api/schedules")
async def api_create_schedule(request: Request,
                              file: UploadFile = File(...),
                              name: str = Form(""),
                              time_str: str = Form(..., alias="time"),
                              interval_days: int = Form(1),
                              needs_screenshot: bool = Form(False)):
    """创建定时采集任务"""
    # 验证时间格式
    try:
        h, m = map(int, time_str.split(":"))
        if not (0 <= h <= 23 and 0 <= m <= 59):
            raise ValueError
    except ValueError:
        raise HTTPException(400, "时间格式错误，应为 HH:MM")

    if interval_days < 1:
        raise HTTPException(400, "间隔天数至少为 1")

    # 保存 ASIN 文件
    os.makedirs(_SCHEDULES_DIR, exist_ok=True)
    import uuid
    sched_id = f"sched_{uuid.uuid4().hex[:8]}"
    ext = os.path.splitext(file.filename or "")[1] or ".txt"
    source_file = os.path.join(_SCHEDULES_DIR, f"{sched_id}{ext}")
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"文件过大：{len(content)//1024//1024}MB，上限 {MAX_UPLOAD_BYTES//1024//1024}MB")
    with open(source_file, "wb") as f:
        f.write(content)

    # 验证文件中有 ASIN
    asin_list = _extract_asins_from_file(source_file)
    if not asin_list:
        os.remove(source_file)
        raise HTTPException(400, "文件中未找到有效 ASIN")

    # 首次创建：last_run_date 设为昨天，确保首次检查时立即触发
    from datetime import timedelta
    yesterday = (_cn_now() - timedelta(days=interval_days)).strftime("%Y-%m-%d")

    sched = {
        "id": sched_id,
        "name": name or f"定时任务-{time_str}",
        "time": time_str,
        "interval_days": interval_days,
        "source_file": source_file,
        "asin_count": len(asin_list),
        "needs_screenshot": needs_screenshot,
        "enabled": True,
        "last_run_date": yesterday,
        "created_at": now_ts(),
    }

    schedules = _get_schedules()
    schedules.append(sched)
    _save_schedules(schedules)

    return {"ok": True, "schedule": sched, "schedules": schedules}


@app.put("/api/schedules/{sched_id}")
async def api_update_schedule(sched_id: str, request: Request):
    """修改定时任务（enabled/time/interval_days/name）"""
    body = await request.json()
    schedules = _get_schedules()
    target = None
    for s in schedules:
        if s.get("id") == sched_id:
            target = s
            break
    if target is None:
        raise HTTPException(404, "定时任务不存在")

    if "enabled" in body:
        target["enabled"] = bool(body["enabled"])
    if "name" in body:
        target["name"] = body["name"]
    if "time" in body:
        try:
            h, m = map(int, body["time"].split(":"))
            if 0 <= h <= 23 and 0 <= m <= 59:
                target["time"] = body["time"]
        except (ValueError, AttributeError):
            pass
    if "interval_days" in body:
        val = int(body["interval_days"])
        if val >= 1:
            target["interval_days"] = val

    _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


@app.delete("/api/schedules/{sched_id}")
async def api_delete_schedule(sched_id: str):
    """删除定时任务 + ASIN 文件"""
    schedules = _get_schedules()
    target = None
    new_schedules = []
    for s in schedules:
        if s.get("id") == sched_id:
            target = s
        else:
            new_schedules.append(s)
    if target is None:
        raise HTTPException(404, "定时任务不存在")

    # 删除 ASIN 文件
    source_file = target.get("source_file", "")
    if source_file and os.path.isfile(source_file):
        os.remove(source_file)

    _save_schedules(new_schedules)
    return {"ok": True, "schedules": new_schedules}


@app.post("/api/schedules/{sched_id}/run")
async def api_run_schedule_now(sched_id: str):
    """手动立即执行一次定时任务"""
    schedules = _get_schedules()
    target = None
    for s in schedules:
        if s.get("id") == sched_id:
            target = s
            break
    if target is None:
        raise HTTPException(404, "定时任务不存在")

    source_file = target.get("source_file", "")
    asin_list = _extract_asins_from_file(source_file)
    if not asin_list:
        raise HTTPException(400, "ASIN 文件为空或不存在")

    now = _cn_now()  # last_run 用中国时间
    # P4.7：**分钟 -> 秒**（有意的行为改动，见 _batch_name 的 docstring）。
    # 这一处和 _auto_scrape_scheduler 那一处正是会互相撞名的两个。
    batch_name = _batch_name(f"auto_{target.get('name', 'task')}")
    zc = _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE)
    ns = target.get("needs_screenshot", False)
    batch_id = await db.create_batch(batch_name, ns, is_auto=True)
    await db.create_tasks(batch_id, asin_list, zc, ns)

    target["last_run_date"] = now.strftime("%Y-%m-%d")
    _save_schedules(schedules)
    logger.info(f"手动执行定时任务: {batch_name}, {len(asin_list)} ASINs")

    return {"ok": True, "batch_id": batch_id, "batch_name": batch_name, "asin_count": len(asin_list)}


# ==================== 兼容旧 schedule API（settings.html 旧 UI 调用） ====================

@app.get("/api/auto-scrape/schedules")
async def api_legacy_list_schedules():
    return {"schedules": _get_schedules()}


@app.post("/api/auto-scrape/schedules")
async def api_legacy_add_schedule(request: Request):
    """旧式简单定时（无文件，使用全库 ASIN）"""
    body = await request.json()
    time_str = body.get("time", "")
    try:
        h, m = map(int, time_str.split(":"))
        if not (0 <= h <= 23 and 0 <= m <= 59):
            raise ValueError
    except ValueError:
        raise HTTPException(400, "时间格式错误")

    import uuid
    sched = {
        "id": f"sched_{uuid.uuid4().hex[:8]}",
        "name": f"全库采集-{time_str}",
        "time": time_str,
        "interval_days": 1,
        "source_file": "",  # 空=全库 ASIN
        "asin_count": 0,
        "needs_screenshot": False,
        "enabled": True,
        "last_run_date": "",
        "created_at": now_ts(),
    }
    schedules = _get_schedules()
    schedules.append(sched)
    _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


@app.put("/api/auto-scrape/schedules/{index}")
async def api_legacy_toggle_schedule(index: int, request: Request):
    body = await request.json()
    schedules = _get_schedules()
    if 0 <= index < len(schedules):
        if "enabled" in body:
            schedules[index]["enabled"] = bool(body["enabled"])
        _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


@app.delete("/api/auto-scrape/schedules/{index}")
async def api_legacy_delete_schedule(index: int):
    schedules = _get_schedules()
    if 0 <= index < len(schedules):
        removed = schedules.pop(index)
        sf = removed.get("source_file", "")
        if sf and os.path.isfile(sf):
            os.remove(sf)
        _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


# ==================== Recon 侦查端点（锁竞争 / 阶段耗时）====================
# 已搬到 server/api/debug.py（Phase 3.2）：_pct / _summary / lock-stats /
# event-stream / lock-stats-reset，连同 DELETE /api/database 的裸事务。
# router 在上面「API: 诊断 / 侦查」那一节挂载。


# ==================== 入口 ====================

def main():
    import uvicorn
    uvicorn.run(
        "server.app:app",
        host=config.SERVER_HOST,
        port=config.SERVER_PORT,
        log_level="info",
    )


if __name__ == "__main__":
    main()
